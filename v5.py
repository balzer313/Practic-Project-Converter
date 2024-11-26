from datetime import datetime
import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import docx
from docx.enum.text import WD_PARAGRAPH_ALIGNMENT
from docx.shared import Pt

import json
import pandas as pd
from PIL import Image, ImageTk
from openpyxl import load_workbook
from openpyxl.styles import numbers
from openpyxl.styles import Font, Alignment


def convert_time(timestamp):
    return datetime.utcfromtimestamp(timestamp / 1000).strftime('%b %d, %Y')
def discount_breakdown(adjustments):
    for adjustment in adjustments:
        if adjustment.get('promoCode') is not None and adjustment.get('type') == 'PROMO_CODE':
            percent = adjustment.get('percent', 0)
            promo_code = adjustment.get('promoCode')
            return f"{percent}% PROMO_CODE: {promo_code}"
    return None
def select_json_file():
    global selected_json_file
    selected_json_file = filedialog.askopenfilename(title="Select Json File",
                                                    filetypes=(("Json Files", "*.json"), ("All Files", "*.*")))
    if not selected_json_file:
        return  # User canceled file selection
    display_selected_file()
def display_selected_file():#screen 2
    global frame2  # Declare new_frame as a global variable
    global generate_word_file
    frame.grid_forget()  # Remove the existing frame
    frame2 = ttk.Frame(root, padding=20)
    frame2.grid(row=0, column=0)

    new_title_label = ttk.Label(frame2, text="Selected Excel File", font=("Arial", 16, "bold"))
    new_title_label.grid(row=0, column=0, pady=(0, 20), columnspan=2)

    excel_file_label = ttk.Label(frame2, text=selected_json_file, font=("Arial", 10))
    excel_file_label.grid(row=1, column=0, pady=(0, 10), padx=10, sticky='w')

    change_excel_button = ttk.Button(frame2, text="Change Json File", command=change_selected_file)
    change_excel_button.grid(row=2, column=0, pady=(10, 10), padx=10, sticky='w')

    generate_word_checkbox = ttk.Checkbutton(frame2, text="Generate Word File", variable=generate_word_file)
    generate_word_checkbox.grid(row=3, column=0, pady=(10, 10), padx=10, sticky='w')

    convert_button = ttk.Button(frame2, text="Convert", command=convert)
    convert_button.grid(row=4, column=0, pady=(10, 10), padx=10, sticky='w')

    # Center all widgets both vertically and horizontally
    frame2.grid_rowconfigure(1, weight=1)
    frame2.grid_columnconfigure(0, weight=1)

def change_selected_file():
    global selected_json_file
    global frame2  # Make new_frame accessible
    selected_json_file = None
    frame2.grid_forget()  # Remove the existing frame
    frame.grid()  # Re-display the original frame to select a new file2

def convert():
    global selected_json_file
    output_folder = filedialog.askdirectory(title="Select Folder to Save File")
    if output_folder:
        try:
            # Load the JSON file
            with open(selected_json_file, 'r') as f:
                data = json.load(f)

            # Check the type of JSON file to determine the processing logic
            if "QuoteDetails" in data:
                # Handle quote JSON file (quote_24005498...json)
                process_quote_file(data, output_folder)
            elif "orderIntentId" in data:
                # Handle estimation JSON file (estimate.json)
                process_estimate_file(data, output_folder)
            else:
                raise ValueError("Unknown JSON file structure")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")


def process_quote_file(data, output_folder):
    # This function processes quote JSON files
    quote_details = data.get('QuoteDetails', {})
    line_items = quote_details.get('upcomingBills', {}).get('lines', [])
    for_users = quote_details.get("lineItems", [])

    table_data = []
    summary_data = {}  # Dictionary to hold aggregated data for duplicate items
    annu_month = "MONTHLY"

    for item in line_items:
        line_item = {}
        product_name = item['description']
        line_item['Product'] = product_name

        if "ANNUAL" in item['description']:
            annu_month = "ANNUAL"

        line_item['List Price'] = item['subTotal'] / 100
        line_item['Discount'] = (item['subTotal'] - item['total'] - (item['tax'] if 'tax' in item else 0)) / 100
        line_item['Amount excl. tax'] = (item['total'] + item['tax'] if 'tax' in item else 0) / 100
        line_item['Tax'] = (item['tax'] if 'tax' in item else 0) / 100
        line_item['Amount (USD)'] = item['total'] / 100
        line_item['Billing period'] = f"{convert_time(item['period']['startsAt'])} - {convert_time(item['period']['endsAt'])}"

        # Determine the correct quantity
        users_quantity = item.get('quantity', 0)
        for item_up in for_users:
            if item_up.get("lineItemId") == item['quoteLineId']:
                try:
                    users_quantity = item_up.get("chargeQuantities", [])[0].get("quantity", users_quantity)
                except Exception:
                    pass
        line_item['Users'] = users_quantity

        line_item['Discounts breakdown'] = discount_breakdown(item.get('adjustments', []))

        not_to_check = ['Product', 'Users', 'Billing period', 'Discounts breakdown']
        all_zero = all(
            float(value) == 0
            for key, value in line_item.items() if key not in not_to_check and isinstance(value, (int, float))
        )
        if all_zero:  # checks if all the line is 0
            continue

        # Add the line item to table_data
        table_data.append(line_item)

        # Accumulate totals for the summary
        if product_name not in summary_data:
            summary_data[product_name] = {
                'Product name': product_name,
                'Users': users_quantity,
                'List price': line_item['List Price'],
                'Discount': line_item['Discount'],
                'Total': line_item['Amount excl. tax'],
            }
        else:
            # Aggregate values for duplicate items
            summary_data[product_name]['List price'] += line_item['List Price']
            summary_data[product_name]['Discount'] += line_item['Discount']
            summary_data[product_name]['Total'] += line_item['Amount excl. tax']

    # Convert the summary data to a list for the summary sheet
    summary_table = list(summary_data.values())
    save_to_excel_and_word(table_data, data, output_folder, annu_month, summary_table)



def process_estimate_file(data, output_folder):
    # Process estimation JSON files and create both detailed and summary data
    items = data.get('items', [])
    table_data = []
    summary_data = {}

    for item in items:
        for sub_item in item.get('items', []):
            # Extract relevant details
            charge_quantity = sub_item.get('chargeQuantity', {}).get('quantity', None)
            product_name = sub_item.get('offeringId', 'Unknown Offering')
            description = f"{product_name}, up to {charge_quantity} users"

            list_price = sub_item.get('subtotal', 0)
            discount = sum(adj.get('amount', 0) for adj in sub_item.get('adjustments', []))
            total = sub_item.get('total', 0)

            # Detailed line item
            line_item = {
                'Product': description,
                'List Price': list_price,
                'Discount': discount,
                'Amount excl. tax': total,
                'Tax': sub_item.get('tax', 0),
                'Amount (USD)': total,
                'Billing period': f"{convert_time(sub_item.get('billPeriod', {}).get('start'))} - {convert_time(sub_item.get('billPeriod', {}).get('end'))}",
                'Users': charge_quantity,
                'Discounts breakdown': discount_breakdown(sub_item.get('adjustments', [])),
            }

            table_data.append(line_item)

            # Accumulate totals for the summary
            if product_name not in summary_data:
                summary_data[product_name] = {
                    'Product name': product_name,
                    'Users': charge_quantity or 0,
                    'List price': list_price,
                    'Discount': discount,
                    'Total': total,
                }
            else:
                # Accumulate values
                summary_data[product_name]['Users'] += charge_quantity or 0
                summary_data[product_name]['List price'] += list_price
                summary_data[product_name]['Discount'] += discount
                summary_data[product_name]['Total'] += total

    # Prepare summary table
    summary_table = list(summary_data.values())

    # Save to Excel and Word with an additional summary sheet
    save_to_excel_and_word(table_data, data, output_folder, annu_month="MONTHLY", summary_table=summary_table)



def save_to_excel_and_word(table_data, data, output_folder, annu_month, summary_table=None):
    if not table_data:
        messagebox.showerror("Error", "No data available to create the DataFrame.")
        return

    # Convert the table data to a DataFrame
    df = pd.DataFrame(table_data)
    df.columns = ['Product', 'List Price (USD)', 'Discount', 'Amount excl. tax (USD)', 'Tax', 'Amount',
                  'Billing period', 'Users', 'Discounts breakdown']

    company_name = data.get('InvoiceGroup', {}).get('shipToParty', {}).get('name', 'Company_Name')
    quote_number = data.get('QuoteDetails', {}).get('number', 'Estimation')

    # Save Excel and generate word file if needed
    excel_name = f'{annu_month} Excel - {company_name}.xlsx'
    output_excel = os.path.join(output_folder, excel_name)

    df.to_excel(output_excel, index=False, engine='openpyxl')

    # Additional formatting logic for Excel
    workbook = load_workbook(output_excel)
    worksheet = workbook.active

    # Define the number format with commas and two decimal places
    number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
    number_format_integer = numbers.FORMAT_NUMBER  # No decimals, just commas
    # make the right columns has the numbers format: 1,234,567(,)
    for col in ['B', 'C', 'D', 'E', 'F']:  # Columns to format
        for cell in worksheet[col][1:]:  # Skip the header row
            if isinstance(cell.value, (int, float)):  # Only format numeric cells
                cell.number_format = number_format
    for cell in worksheet['H'][1:]:  # Skip the header row
        if isinstance(cell.value, (int, float)):  # Only format numeric cells
            cell.number_format = number_format_integer
    # adding length to a line which doesnt has enought space
    for col in worksheet.columns:
        max_length = 0
        column_letter = col[0].column_letter  # Get column letter (e.g., 'A')
        for cell in col:
            if cell.value is not None:
                # Calculate the display length of the formatted value
                if isinstance(cell.value, (int, float)):
                    formatted_value = format(cell.value, ",.2f")
                    max_length = max(max_length, len(formatted_value))
                else:
                    max_length = max(max_length, len(str(cell.value)))

        # Apply the width to the column
        worksheet.column_dimensions[column_letter].width = max_length + 2

    # add the summary_table for duplicate items
    if summary_table:
        # Create the "Summary" sheet
        summary_sheet = workbook.create_sheet(title="Summary")

        # Use the same headers as the regular sheet
        summary_headers = ['Product', 'List Price (USD)', 'Discount', 'Amount excl. tax (USD)', 'Tax',
                           'Amount', 'Billing period', 'Users', 'Discounts breakdown']
        summary_sheet.append(summary_headers)

        for col_num, header in enumerate(summary_headers, start=1):
            cell = summary_sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)  # Make the header bold
            cell.alignment = Alignment(horizontal='center')  # Center-align the header
        # Populate rows for the summary data
        for row in summary_table:
            # Ensure all headers exist in the summary data
            summary_row = [
                row.get('Product name', ''),
                row.get('List price', 0),
                row.get('Discount', 0),
                row.get('Total', 0),  # Amount excl. tax (USD)
                0,  # Tax is assumed as 0 since no tax is provided
                row.get('Total', 0),  # Amount (USD)
                '',  # Billing period is not applicable for summary
                row.get('Users', 0),
                '',  # Discounts breakdown not applicable in summary
            ]
            summary_sheet.append(summary_row)

        # Apply formatting to the "Summary" sheet
        for col in ['B', 'C', 'D', 'E', 'F', 'H']:  # Columns to format
            for cell in summary_sheet[col][1:]:  # Skip the header row
                if isinstance(cell.value, (int, float)):  # Only format numeric cells
                    cell.number_format = number_format

        # Adjust column widths in the "Summary" sheet
        for col in summary_sheet.columns:
            max_length = 0
            column_letter = col[0].column_letter  # Get column letter (e.g., 'A')
            for cell in col:
                if cell.value is not None:
                    # Calculate the display length of the formatted value
                    if isinstance(cell.value, (int, float)):
                        formatted_value = format(cell.value, ",.2f")
                        max_length = max(max_length, len(formatted_value))
                    else:
                        max_length = max(max_length, len(str(cell.value)))

            # Apply the width to the column
            summary_sheet.column_dimensions[column_letter].width = max_length + 2

    workbook.save(output_excel)  # save changes

    # Optionally generate Word file
    word_file()  # Assuming word_file() checks `generate_word_file` before generating
    current_time = datetime.now().strftime('%Y-%m-%d, %H:%M:%S')  # Get the current date and time
    messagebox.showinfo("Success",
                        f"Excel file created at {output_excel}\nCompany: {company_name}\nQuote number: {quote_number}\nUpdated: {current_time}")
    root.destroy()


def word_file(): ## creating(if wonted) a word file
    global generate_word_file
    def change_word(docs, old_word, new_word, bold, size, page_break):
        for paragraph in docs.paragraphs:
            if f'%%{old_word}%%' in paragraph.text:
                # print('ok')
                paragraph.text = paragraph.text.replace(f'%%{old_word}%%',
                                                        new_word)  # Replace old_word with the new_word
                for run in paragraph.runs:
                    if new_word in run.text:
                        run.font.size = Pt(size)  # Set font size to 18pt
                        run.font.bold = bold  # Set text to bold
                        run.alignment = WD_PARAGRAPH_ALIGNMENT.CENTER  # Center align the text
                if page_break:
                    paragraph.clear()
                    run2 = paragraph.add_run()
                    run2.add_break(docx.enum.text.WD_BREAK.PAGE)
    if generate_word_file:
        return ## handle the word file

money_type = "DOLLAR"

def toggle_currency():
    global money_type
    if money_type == "DOLLAR": money_type = "DIRHAM"; toggle_currency_button.configure(image=off_icon)
    else: money_type = "DOLLAR"; toggle_currency_button.configure(image=on_icon)


# Set up the Tkinter window
root = tk.Tk()
root.title("JSON to Excel and Docx Converter")

# Get the screen width and height
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
window_width = 600
window_height = 300
x = (screen_width - window_width) // 2
y = (screen_height - window_height) // 2
root.geometry(f"{window_width}x{window_height}+{x}+{y}")

# Load images for on and off states
on_image = Image.open("images/usa_flag.png").resize((45, 30))
on_icon = ImageTk.PhotoImage(on_image)
off_image = Image.open("images/Flag-United-Arab-Emirates.png").resize((45, 30))
off_icon = ImageTk.PhotoImage(off_image)

currency_var = tk.BooleanVar(value=True)  # True for DOLLAR, False for DIRHAM
generate_word_file = tk.BooleanVar(value=False)

# Create a frame with padding
frame = ttk.Frame(root, padding=20)
frame.grid(row=0, column=0, sticky="nsew")

# Create a label for the title
title_label = ttk.Label(frame, text="JSON to Excel and Docx Converter", font=("Arial", 16, "bold"))
title_label.grid(row=0, column=0, pady=(0, 20), columnspan=2)

# Create a button to select the JSON file
select_excel_button = ttk.Button(frame, text="Select JSON File", command=select_json_file)
select_excel_button.grid(row=1, column=0, pady=(0, 10), padx=10, sticky='w')

# Label to display the selected file name
json_file_label = ttk.Label(frame, text="", font=("Arial", 10))
json_file_label.grid(row=2, column=0, pady=(10, 10), padx=10, sticky='w')

# Currency toggle button
toggle_currency_button = ttk.Checkbutton(
    frame, image=on_icon, variable=currency_var, command=toggle_currency,
    style="TButton", onvalue=True, offvalue=False
)
toggle_currency_button.grid(row=3, column=0, pady=(0, 10), padx=10, sticky='w')

# Configure window grid to make it responsive
root.grid_rowconfigure(0, weight=1)
root.grid_columnconfigure(0, weight=1)

root.mainloop()
